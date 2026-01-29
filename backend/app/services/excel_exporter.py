"""Excel export service — modifies the original uploaded file in-place.

Opens the original Excel file (preserving formatting, extra sheets, column
positions, everything) and ONLY overwrites the Product Name and Description
cells for approved product groups.  Matching is by (product_name,
product_token, sku) — the same composite key used for variant grouping —
so every variant row in a group gets the same generated content.
"""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


class ExcelExporter:
    """Export by patching the original uploaded Excel file."""

    # Exact header strings from Faire templates
    PRODUCT_NAME_HEADER = "Product Name (English)"
    DESCRIPTION_HEADER = "Description (English)"
    PRODUCT_TOKEN_HEADER = "Product Token"
    SKU_HEADER = "SKU"

    def export(
        self,
        original_file_path: Path | str,
        groups_lookup: dict[tuple[str, str, str], dict],
        include_pending: bool = False,
    ) -> BytesIO:
        """Patch the original Excel with generated content for approved groups.

        Args:
            original_file_path: Path to the original uploaded .xlsx file.
            groups_lookup: Dict keyed by (product_name, product_token, sku)
                with values containing:
                  - review_status, status (group generation status)
                  - generated_title, generated_description
                  - edited_title, edited_description
            include_pending: Also update content for generated-but-unreviewed
                products (not rejected).

        Returns:
            BytesIO buffer containing the patched .xlsx file.
        """
        # Load original workbook preserving formatting, styles, extra sheets
        wb = load_workbook(str(original_file_path))

        # Find the data sheet (Faire uses "Products" sheet name)
        if "Products" in wb.sheetnames:
            ws = wb["Products"]
        else:
            ws = wb.active

        # Locate required columns from the header row (row 1)
        col_indices = self._find_columns(ws)
        if not col_indices:
            # If we can't find the columns, return the file unmodified
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            wb.close()
            return buffer

        name_col, desc_col, token_col, sku_col = col_indices

        # Walk every data row and patch approved products
        for row_idx in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=name_col).value
            cell_token = ws.cell(row=row_idx, column=token_col).value
            cell_sku = ws.cell(row=row_idx, column=sku_col).value

            # Skip empty / header-repeat rows
            if not cell_token or not cell_sku:
                continue

            # Build the same composite key the DB uses.
            # Apply the same formula-injection sanitisation the upload parser
            # applies so the key matches the stored value.
            key = (
                self._sanitize(cell_name) if cell_name else "",
                str(cell_token),
                str(cell_sku),
            )

            group = groups_lookup.get(key)
            if group is None:
                continue  # row not in our database

            if not self._should_update(group, include_pending):
                continue  # not approved, keep original values

            # Compute effective title / description
            effective_title = (
                group.get("edited_title")
                or group.get("generated_title")
                or cell_name  # fallback: leave original
            )
            effective_desc = (
                group.get("edited_description")
                or group.get("generated_description")
                or ws.cell(row=row_idx, column=desc_col).value
            )

            ws.cell(row=row_idx, column=name_col).value = effective_title
            ws.cell(row=row_idx, column=desc_col).value = effective_desc

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()
        return buffer

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------

    def _find_columns(self, ws) -> tuple[int, int, int, int] | None:
        """Return (name_col, desc_col, token_col, sku_col) 1-indexed, or None."""
        name_col = desc_col = token_col = sku_col = None

        for col_idx, cell in enumerate(ws[1], 1):
            header = cell.value
            if header == self.PRODUCT_NAME_HEADER:
                name_col = col_idx
            elif header == self.DESCRIPTION_HEADER:
                desc_col = col_idx
            elif header == self.PRODUCT_TOKEN_HEADER:
                token_col = col_idx
            elif header == self.SKU_HEADER:
                sku_col = col_idx

        if all(v is not None for v in (name_col, desc_col, token_col, sku_col)):
            return name_col, desc_col, token_col, sku_col
        return None

    @staticmethod
    def _sanitize(value) -> str:
        """Mirror the upload parser's formula-injection sanitisation."""
        s = str(value) if value is not None else ""
        if s and s[0] in ("=", "+", "-", "@"):
            return "'" + s
        return s

    @staticmethod
    def _should_update(group: dict, include_pending: bool) -> bool:
        """Decide whether a group's rows should receive generated content."""
        review_status = group.get("review_status")
        gen_status = group.get("status")

        if review_status in ("approved", "edited"):
            return True

        if include_pending and gen_status == "generated" and review_status != "rejected":
            return True

        return False
