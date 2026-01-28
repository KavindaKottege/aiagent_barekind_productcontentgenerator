"""Streaming Excel parser using openpyxl read_only mode."""
from pathlib import Path
from openpyxl import load_workbook
from typing import Generator


class ExcelParser:
    """Parse Excel files in streaming mode for memory efficiency."""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self.wb = None

    def parse(self) -> Generator[list[dict], None, None]:
        """
        Stream Excel rows in batches.
        Yields batches of 500 rows as list[dict].
        IMPORTANT: Caller must call close() when done.
        """
        self.wb = load_workbook(filename=str(self.file_path), read_only=True)

        # Look for "Products" sheet (common in Faire templates)
        # If not found, use the active sheet
        if 'Products' in self.wb.sheetnames:
            ws = self.wb['Products']
        else:
            ws = self.wb.active

        # Get headers from first row
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(cell) if cell else f"column_{i}" for i, cell in enumerate(next(rows_iter))]

        batch = []
        for row in rows_iter:
            # Skip completely empty rows
            if not any(row):
                continue

            row_dict = dict(zip(headers, row))

            # Skip invalid metadata/header rows (common in Faire templates)
            # These rows have placeholder values like "info_", "Optional", "product_name", etc.
            if self._is_metadata_row(row_dict):
                continue

            batch.append(row_dict)

            if len(batch) >= 500:
                yield batch
                batch = []

        if batch:
            yield batch

    def _is_metadata_row(self, row_dict: dict) -> bool:
        """Check if a row is a metadata/header row that should be skipped."""
        # Common indicator columns in Faire templates
        indicator_cols = ['SKU', 'Product Token', 'Product Name', 'Status']

        # Get first available indicator value
        indicator_value = None
        for col in indicator_cols:
            if col in row_dict and row_dict[col]:
                indicator_value = str(row_dict[col]).strip().lower()
                break

        if not indicator_value:
            return True  # Empty indicator = skip

        # Skip rows with placeholder/metadata patterns
        invalid_patterns = [
            'info_',           # info_product_type, info_status, etc.
            'optional',        # Optional placeholder
            'sku',             # Header row repeated
            'product_token',   # Header row repeated
            'product_name',    # Header row repeated
            'status',          # Header row repeated
            'example',         # Example data
            'sample',          # Sample data
        ]

        for pattern in invalid_patterns:
            if pattern in indicator_value:
                return True

        return False

    def close(self):
        """Close workbook. MUST be called after parsing."""
        if self.wb:
            self.wb.close()
            self.wb = None
